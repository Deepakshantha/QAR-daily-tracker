import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from fuzzywuzzy import process

# ══════════════════════════════════════════════════════════════════════════════
#  LOOKUP LISTS
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_NAMES = [
    "Abarajithan Govindarajan",
    "Aishwarya Rajamohan",
    "Akash P M",
    "Akash.M Murali",
    "Akshay Kumar P",
    "Akshay V Kumar",
    "Allan Augustine",
    "Anshuman Dey",
    "Anusree Anil",
    "Archana Venkatesan",
    "Ashwin Kumar",
    "Avi Sharma",
    "Ayyapparaj Dhamodharan",
    "Bala Thirupathi Raaja",
    "Barathi Priya",
    "Bhavani Dhanabalan",
    "Deepika Raghuraj",
    "Deepika Subramani",
    "Devanathan",
    "Dilip Suresh",
    "Divya Dharshini",
    "Divya Shree",
    "Durairaj Saravanakumar",
    "GN Karthik",
    "Gnana Jenifer Wilciya",
    "Gurumoorthy Vijayarangan",
    "Hemavathy R",
    "Harshaavardhan Subramani",
    "V Jahnavi",
    "Jeff Rohit",
    "Joshni N",
    "Jenithson Thommai",
    "Karthick Gurunathan",
    "Karthik A",
    "Karthikeyan Panachavaranam",
    "Keerthana B",
    "Kiranraj Ravichandran",
    "Kishore Saravanan",
    "Kumaran Ramachandran",
    "Maha S",
    "Mariya Antony Britto",
    "Md Shadman Hayat Siddhiquie",
    "Meenakshi Maragathavel",
    "Melwin Manoj",
    "Mohammed Wihaj",
    "Moneshwar Devaraj",
    "Mukul Vyas Parameswar",
    "Nadhiya Siva Subramanian",
    "Najir Hussain Nashim Miyan",
    "Neelufur Begam",
    "Nishanthini Umapathy",
    "Palani Raja Vellaisamy",
    "Parthasarathy Letchumanan",
    "Prabhakaran Sekar",
    "Pragadeeshwaran Ganesan",
    "Priya Dharshini K",
    "Priyea Dharshani B",
    "Ritesh Suresh",
    "Rohit Subramani",
    "Rojini.S Sathish Kumar",
    "Rex Fleming",
    "Sabariraj Iyyappan",
    "Sachin Rajesh",
    "Samyuktha Balakrishnaian",
    "Saquib Tanweer",
    "Sarathirajan K",
    "SarathKumar Ravikumar",
    "Sathish Kumar Venkatesan",
    "Shalini S",
    "Shalini Subramanian",
    "Shantha Kumar Saravanan",
    "Sivasankari Arumugam",
    "Sonia Selva Kumar",
    "Sridevi Rangarajan",
    "Sruthi Mathivanan",
    "Steffin T M",
    "Saranesh Duraisamy",
    "Swetha Mani",
    "Tamilarasi Balamurugan",
    "Veera Sabarinathan",
    "Vignesh Murugan",
    "Vijay Kumar R",
    "Vijayalakshmi Dhanabalan",
    "Vijayalakshmi Janakiraman",
    "Vinod Ram",
    "Vishwa Alagiri",
    "Yazhini Krishnamoorthy",
    "Yuvaraj Selvam"
]

DEFAULT_REPORT_NAMES = [
    "Missing Copy Report", "DEX Advanced Find Report",
    "Missing Script Report", "SET Discrepancy Report",
    "Posting", "Post-Log", "Pre-Log",
]

DEFAULT_MARKETS_REGIONS = [
    "Atlantic", "Northeast", "Midwest", "Southeast", "Central", "Pacific", "All",
]

DEFAULT_STATUSES = [
    "Completed", "Post Available", "CM Action", "No Action",
]

EXPECTED_COLUMNS = [
    "Date", "Month", "Name", "Report Name", "Markets / Regions",
    "Contract", "Status", "CM Action Reason", "TAT Miss", "Quality Met",
    "Error Type", "Error Comments", "Revenue Impact", "Feedback Delivered by",
    "Year",
]

FUZZY_THRESHOLD = 80


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def is_empty(val):
    return pd.isna(val) or str(val).strip() == ""


def fuzzy_match(value, valid_list, threshold=FUZZY_THRESHOLD):
    if is_empty(value) or not valid_list:
        return str(value).strip() if not is_empty(value) else "", False
    val_str = str(value).strip()
    match, score = process.extractOne(val_str, valid_list)
    return (match, True) if score >= threshold else (val_str, False)


def get_date_info():
    today    = datetime.today()
    target   = today - timedelta(days=1)
    date_str = target.strftime("%d-%b-%Y")
    year     = target.year
    if today.day == 1:
        prev      = today.replace(day=1) - timedelta(days=1)
        month_str = prev.strftime("%B")
    else:
        month_str = target.strftime("%B")
    return date_str, month_str, year


# ══════════════════════════════════════════════════════════════════════════════
#  CORE CLEANING
# ══════════════════════════════════════════════════════════════════════════════

def clean_tracker(raw_df, valid_names, valid_report_names,
                  valid_regions, valid_statuses, threshold=FUZZY_THRESHOLD):
    date_str, month_str, current_year = get_date_info()
    raw_df.columns = [str(c).strip() for c in raw_df.columns]
    for col in EXPECTED_COLUMNS:
        if col not in raw_df.columns:
            raw_df[col] = None
    df   = raw_df[EXPECTED_COLUMNS].copy()
    drop = pd.Series(False, index=df.index)

    # 1. DATE
    df["Date"] = date_str

    # 2. MONTH
    df["Month"] = month_str

    # 3. NAME — empty/unknown → drop
    def process_name(val):
        if is_empty(val): return None, True
        matched, good = fuzzy_match(val, valid_names, threshold)
        return (matched, False) if good else (None, True)
    name_res   = df["Name"].apply(process_name)
    df["Name"] = name_res.apply(lambda x: x[0])
    drop      |= name_res.apply(lambda x: x[1])

    # 4. REPORT NAME — empty/unknown → drop
    def process_report(val):
        if is_empty(val): return None, True
        matched, good = fuzzy_match(val, valid_report_names, threshold)
        return (matched, False) if good else (None, True)
    report_res        = df["Report Name"].apply(process_report)
    df["Report Name"] = report_res.apply(lambda x: x[0])
    drop             |= report_res.apply(lambda x: x[1])

    # 5. MARKETS / REGIONS — empty/unknown → drop
    def process_region(val):
        if is_empty(val): return None, True
        matched, good = fuzzy_match(val, valid_regions, threshold)
        return (matched, False) if good else (None, True)
    region_res              = df["Markets / Regions"].apply(process_region)
    df["Markets / Regions"] = region_res.apply(lambda x: x[0])
    drop                   |= region_res.apply(lambda x: x[1])

    # 6. CONTRACT — leave as-is; empty → drop
    def process_contract(val):
        if is_empty(val): return None, True      # empty → drop
        return str(val).strip(), False           # everything else → leave as-is
    contract_res   = df["Contract"].apply(process_contract)
    df["Contract"] = contract_res.apply(lambda x: x[0])
    drop          |= contract_res.apply(lambda x: x[1])

    # 7. STATUS — empty/unknown → drop
    def process_status(val):
        if is_empty(val): return None, True
        matched, good = fuzzy_match(val, valid_statuses, threshold)
        return (matched, False) if good else (None, True)
    status_res   = df["Status"].apply(process_status)
    df["Status"] = status_res.apply(lambda x: x[0])
    drop        |= status_res.apply(lambda x: x[1])

    # 8. CM ACTION REASON
    df["CM Action Reason"] = df["CM Action Reason"].apply(
        lambda v: "-" if is_empty(v) else str(v).strip()
    )

    # 9-12. FIXED VALUES
    df["TAT Miss"]              = "No"
    df["Quality Met"]           = "Yes"
    df["Error Type"]            = "-"
    df["Error Comments"]        = "-"
    df["Revenue Impact"]        = "-"
    df["Feedback Delivered by"] = "-"
    df["Year"]                  = current_year

    # 13. DROP flagged rows
    dropped = int(drop.sum())
    df      = df[~drop].reset_index(drop=True)
    return df, dropped


# ══════════════════════════════════════════════════════════════════════════════
#  PIVOT TABLE
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot(df):
    pivot = (
        df.groupby("Name", as_index=False)
        .agg(**{"Count of Contract": ("Contract", "count")})
        .sort_values("Name").reset_index(drop=True)
    )
    grand = pd.DataFrame([{
        "Name": "Grand Total",
        "Count of Contract": pivot["Count of Contract"].sum(),
    }])
    return pd.concat([pivot, grand], ignore_index=True)


# ══════════════════════════════════════════════════════════════════════════════
#  CHART DATA
# ══════════════════════════════════════════════════════════════════════════════

def build_region_chart_data(df):
    d = df[~df["Markets / Regions"].isin(["-", None])].copy()
    return (d.groupby("Markets / Regions", as_index=False)
             .agg(**{"Contract Count": ("Contract", "count")})
             .sort_values("Contract Count", ascending=False).reset_index(drop=True))

def build_report_chart_data(df):
    d = df[~df["Report Name"].isin(["-", None])].copy()
    return (d.groupby("Report Name", as_index=False)
             .agg(**{"Count": ("Report Name", "count")})
             .sort_values("Count", ascending=False).reset_index(drop=True))

def build_status_chart_data(df):
    d = df[~df["Status"].isin(["-", None])].copy()
    return (d.groupby("Status", as_index=False)
             .agg(**{"Count": ("Status", "count")})
             .sort_values("Count", ascending=False).reset_index(drop=True))


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT — 3 sheets
# ══════════════════════════════════════════════════════════════════════════════

def to_excel_bytes(df, pivot, tracker_sheet_name):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint as DP
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.legend import Legend

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        thin   = Side(style="thin", color="C8D0E0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Sheet 1: Cleaned Data
        df.to_excel(writer, index=False, sheet_name=tracker_sheet_name)
        ws = writer.sheets[tracker_sheet_name]
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        alt_fill = PatternFill("solid", fgColor="EEF2FF")
        wht_fill = PatternFill("solid", fgColor="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            bf = alt_fill if row_idx % 2 == 0 else wht_fill
            for cell in row:
                cell.fill = bf; cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        for cc in ws.columns:
            mx = max((len(str(c.value)) if c.value else 0 for c in cc), default=10)
            ws.column_dimensions[get_column_letter(cc[0].column)].width = min(mx+4, 32)
        ws.row_dimensions[1].height = 34
        ws.freeze_panes = "A2"

        # Sheet 2: Pivot
        pivot.to_excel(writer, index=False, sheet_name="Pivot")
        wp = writer.sheets["Pivot"]
        ph_fill = PatternFill("solid", fgColor="1F3864")
        ph_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        tf      = PatternFill("solid", fgColor="2E4057")
        tfont   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        pa      = PatternFill("solid", fgColor="EEF2FF")
        pw      = PatternFill("solid", fgColor="FFFFFF")
        for cell in wp[1]:
            cell.fill = ph_fill; cell.font = ph_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        tr = len(pivot) + 1
        for row_idx, row in enumerate(wp.iter_rows(min_row=2), start=2):
            is_t = (row_idx == tr + 1)
            bf   = tf if is_t else (pa if row_idx % 2 == 0 else pw)
            for cell in row:
                cell.fill = bf; cell.font = tfont if is_t else Font(name="Calibri", size=10)
                cell.alignment = Alignment(
                    horizontal="left" if cell.column == 1 else "center", vertical="center")
                cell.border = border
        for cc in wp.columns:
            mx = max((len(str(c.value)) if c.value else 0 for c in cc), default=10)
            wp.column_dimensions[get_column_letter(cc[0].column)].width = min(mx+4, 35)
        wp.row_dimensions[1].height = 30
        wp.freeze_panes = "A2"

        # Sheet 3: Charts
        wb = writer.book
        wc = wb.create_sheet("Charts")
        th_fill = PatternFill("solid", fgColor="1F3864")
        th_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        tb_font = Font(name="Calibri", size=10)
        ta      = PatternFill("solid", fgColor="EEF2FF")
        tw      = PatternFill("solid", fgColor="FFFFFF")

        RC = ["2196F3","F44336","8BC34A","9C27B0","00BCD4","FF9800","E91E63"]
        PC = ["E63946","2196F3","2A9D8F","E9C46A","9C27B0","FF9800","00BCD4","8BC34A"]
        SC = ["2196F3","F44336","8BC34A","9C27B0","FF9800","00BCD4"]

        def dl(series):
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.showLegendKey = series.dLbls.showCatName = False
            series.dLbls.showSerName = series.dLbls.showPercent = False

        def sc(ch, title, xt, yt, horiz=False):
            ch.title = title; ch.style = 26; ch.width = 22; ch.height = 15
            if horiz:
                ch.x_axis.title = xt; ch.y_axis.title = yt
                ch.x_axis.majorGridlines = ChartLines(); ch.y_axis.majorGridlines = None
            else:
                ch.x_axis.title = xt; ch.y_axis.title = yt
                ch.y_axis.majorGridlines = ChartLines(); ch.x_axis.majorGridlines = None
            ch.x_axis.numFmt = ch.y_axis.numFmt = "General"
            ch.legend = Legend(); ch.legend.position = "b"

        def wt(wr, sr, scl, hdrs, rows, title):
            tc = wr.cell(row=sr, column=scl, value=title)
            tc.font = Font(bold=True, color="1F3864", name="Calibri", size=11)
            tc.alignment = Alignment(horizontal="left", vertical="center")
            wr.merge_cells(start_row=sr, start_column=scl,
                           end_row=sr, end_column=scl+len(hdrs)-1)
            wr.row_dimensions[sr].height = 20
            hr = sr + 1
            for ci, h in enumerate(hdrs, start=scl):
                c = wr.cell(row=hr, column=ci, value=h)
                c.fill = th_fill; c.font = th_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            wr.row_dimensions[hr].height = 18
            for ri, rd in enumerate(rows):
                dr = hr + 1 + ri
                fl = ta if ri % 2 == 0 else tw
                for ci, val in enumerate(rd, start=scl):
                    c = wr.cell(row=dr, column=ci, value=val)
                    c.fill = fl; c.font = tb_font
                    c.alignment = Alignment(
                        horizontal="left" if ci == scl else "center", vertical="center")
                    c.border = border
                wr.row_dimensions[dr].height = 16
            er = hr + len(rows)
            for ci, h in enumerate(hdrs, start=scl):
                cl = get_column_letter(ci)
                cw = max(len(str(v)) for v in [h]+[str(r[ci-scl]) for r in rows])
                if wr.column_dimensions[cl].width < cw + 4:
                    wr.column_dimensions[cl].width = min(cw+4, 30)
            return hr, er, scl

        # Chart 1 — Region
        rd = build_region_chart_data(df)
        rr = [(r["Markets / Regions"], int(r["Contract Count"])) for _, r in rd.iterrows()]
        rh, re, rs = wt(wc, 1, 1, ["Markets / Regions","Contract Count"], rr, "Markets / Regions Summary")
        c1 = BarChart(); c1.type = "col"; c1.grouping = "clustered"
        sc(c1, "Markets / Regions — Contract Count", "REGION", "CONTRACT COUNT", False)
        c1.add_data(Reference(wc, min_col=rs+1, min_row=rh, max_row=re), titles_from_data=True)
        c1.set_categories(Reference(wc, min_col=rs, min_row=rh+1, max_row=re))
        for si, color in enumerate(RC[:len(rr)]):
            pt = DP(idx=si); pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color; c1.series[0].dPt.append(pt)
        dl(c1.series[0]); wc.add_chart(c1, "E1")

        # Chart 2 — Report Name
        pd2 = build_report_chart_data(df); ps = re + 3
        pr = [(r["Report Name"], int(r["Count"])) for _, r in pd2.iterrows()]
        ph, pe, psc = wt(wc, ps, 1, ["Report Name","Count"], pr, "Report Name Summary")
        c2 = BarChart(); c2.type = "bar"; c2.grouping = "clustered"
        sc(c2, "Report Name — Count Distribution", "COUNT", "REPORT NAME", True)
        c2.add_data(Reference(wc, min_col=psc+1, min_row=ph, max_row=pe), titles_from_data=True)
        c2.set_categories(Reference(wc, min_col=psc, min_row=ph+1, max_row=pe))
        for si, color in enumerate(PC[:len(pr)]):
            pt = DP(idx=si); pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color; c2.series[0].dPt.append(pt)
        dl(c2.series[0]); wc.add_chart(c2, "E26")

        # Chart 3 — Status
        sd = build_status_chart_data(df); ss = pe + 3
        sr2 = [(r["Status"], int(r["Count"])) for _, r in sd.iterrows()]
        sh, se, ssc = wt(wc, ss, 1, ["Status","Count"], sr2, "Status Summary")
        c3 = BarChart(); c3.type = "col"; c3.grouping = "clustered"
        sc(c3, "Status — Count Distribution", "STATUS", "COUNT", False)
        c3.add_data(Reference(wc, min_col=ssc+1, min_row=sh, max_row=se), titles_from_data=True)
        c3.set_categories(Reference(wc, min_col=ssc, min_row=sh+1, max_row=se))
        for si, color in enumerate(SC[:len(sr2)]):
            pt = DP(idx=si); pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color; c3.series[0].dPt.append(pt)
        dl(c3.series[0]); wc.add_chart(c3, "E51")

        wc.sheet_view.showGridLines = False
        wc.sheet_properties.tabColor = "1F3864"

    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="QAR Daily Tracker Cleaner", page_icon="📋", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');
html, body, * { font-family: 'Inter', sans-serif; }
[data-testid="stAppViewContainer"] { background: #0B0F1A; }
[data-testid="stHeader"]           { background: transparent; }
section[data-testid="stSidebar"]   { background: #0D1120; }
.hero { margin-bottom: 2rem; }
.hero-badge {
    display: inline-block; background: #151C30; border: 1px solid #1A3040;
    border-radius: 100px; padding: 3px 14px; font-size: 0.7rem;
    font-family: 'Space Mono', monospace; color: #22C55E;
    letter-spacing: 1.5px; margin-bottom: 14px;
}
.hero-title {
    font-family: 'Space Mono', monospace; font-size: 1.85rem; font-weight: 700;
    color: #E4EAFF; line-height: 1.2; margin-bottom: 6px;
}
.hero-sub { font-size: 0.92rem; color: #5A6A8A; margin: 0; }
.card { background: #111827; border: 1px solid #1C2640; border-radius: 14px; padding: 24px 28px; margin-bottom: 18px; }
.card-accent { border-left: 4px solid #22C55E; }
.step-num { font-size: 0.65rem; font-family: 'Space Mono', monospace; color: #22C55E; letter-spacing: 2px; text-transform: uppercase; margin-bottom: 6px; }
.card-title { font-size: 1rem; font-weight: 600; color: #C5D0F0; margin-bottom: 4px; }
.card-desc  { font-size: 0.82rem; color: #4A5878; line-height: 1.55; }
.rules-wrap { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 10px; }
.rule-pill { background: #131A2E; border: 1px solid #1A3040; border-radius: 8px; padding: 9px 12px; font-size: 0.76rem; color: #6A7FA8; line-height: 1.45; }
.rule-pill b { color: #B8C8F0; display: block; margin-bottom: 2px; }
.file-card { background: #0E1624; border: 1px solid #1A2640; border-radius: 10px; padding: 16px 20px; margin: 12px 0; display: flex; align-items: center; gap: 12px; }
.file-icon { font-size: 1.5rem; }
.file-meta { flex: 1; }
.file-name   { font-weight: 600; color: #C5D0F0; font-size: 0.9rem; }
.file-detail { color: #3D5070; font-size: 0.78rem; margin-top: 2px; }
.success-card { background: linear-gradient(135deg, #091A10 0%, #0C2015 100%); border: 1px solid #164D28; border-radius: 12px; padding: 18px 22px; color: #34D399; font-family: 'Space Mono', monospace; font-size: 0.88rem; margin: 14px 0; }
.warn-note { color: #FBBF24; font-size: 0.8rem; margin-top: 6px; font-family: 'Inter', sans-serif; }
.hdiv { border: none; border-top: 1px solid #151E30; margin: 22px 0; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <div class="hero-badge">📋 QAR OPS · INTERNAL TOOL</div>
  <div class="hero-title">QAR Daily Tracker Cleaner</div>
  <p class="hero-sub">Upload the raw messy QAR tracker → get a clean Excel with pivot & charts.</p>
</div>
""", unsafe_allow_html=True)

date_str, month_str, current_year = get_date_info()
output_filename  = f"QAR Daily Tracker ({date_str}).xlsx"
tracker_tab_name = f"QAR Daily Tracker {date_str}"

st.markdown("""
<div class="card card-accent">
  <div class="step-num">Step 01</div>
  <div class="card-title">Upload Your Messy Tracker</div>
  <div class="card-desc">Select the unclean Excel file (.xlsx or .xls) from your computer.</div>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("", type=["xlsx", "xls"], label_visibility="collapsed")

with st.expander("📋 Cleaning rules reference"):
    st.markdown("""
    <div class="rules-wrap">
      <div class="rule-pill"><b>Date</b>Always → current date − 1 (DD-Mon-YYYY)</div>
      <div class="rule-pill"><b>Month</b>Current month; if today = 1st → previous month</div>
      <div class="rule-pill"><b>Name</b>Fuzzy-matched · empty or unknown → row removed</div>
      <div class="rule-pill"><b>Report Name</b>Fuzzy-matched · empty or unknown → row removed</div>
      <div class="rule-pill"><b>Markets / Regions</b>Fuzzy-matched · empty or unknown → row removed</div>
      <div class="rule-pill"><b>Contract</b>Left as-is · empty → row removed</div>
      <div class="rule-pill"><b>Status</b>Fuzzy-matched · empty or unknown → row removed</div>
      <div class="rule-pill"><b>CM Action Reason</b>Left as-is · empty → "-"</div>
      <div class="rule-pill"><b>TAT Miss</b>Always → No</div>
      <div class="rule-pill"><b>Quality Met</b>Always → Yes</div>
      <div class="rule-pill"><b>Error Type / Comments</b>Always → -</div>
      <div class="rule-pill"><b>Revenue Impact</b>Always → -</div>
      <div class="rule-pill"><b>Feedback Delivered by</b>Always → -</div>
      <div class="rule-pill"><b>Year</b>Always → current year</div>
      <div class="rule-pill"><b>Pivot Sheet</b>Name · Count of Contract · Grand Total</div>
      <div class="rule-pill"><b>Charts Sheet</b>Region · Report Name · Status counts (style 26)</div>
    </div>
    """, unsafe_allow_html=True)

with st.expander("⚙️ Customise lookup lists & sensitivity"):
    c1, c2 = st.columns(2)
    with c1:
        names_txt   = st.text_area("✅ Valid Names (one per line)",        value="\n".join(DEFAULT_NAMES),           height=200)
        regions_txt = st.text_area("✅ Valid Markets / Regions (one per line)", value="\n".join(DEFAULT_MARKETS_REGIONS), height=140)
    with c2:
        reports_txt = st.text_area("✅ Valid Report Names (one per line)", value="\n".join(DEFAULT_REPORT_NAMES),    height=160)
        status_txt  = st.text_area("✅ Valid Statuses (one per line)",     value="\n".join(DEFAULT_STATUSES),        height=160)
    sensitivity = st.slider("Fuzzy match sensitivity (%)", 50, 100, FUZZY_THRESHOLD)

def parse_list(txt):
    return [l.strip() for l in txt.split("\n") if l.strip()]

valid_names   = parse_list(names_txt)
valid_regions = parse_list(regions_txt)
valid_reports = parse_list(reports_txt)
valid_status  = parse_list(status_txt)

if uploaded_file:
    try:
        raw_df = pd.read_excel(uploaded_file)
        st.markdown(f"""
        <div class="file-card">
          <div class="file-icon">📂</div>
          <div class="file-meta">
            <div class="file-name">{uploaded_file.name}</div>
            <div class="file-detail">{len(raw_df):,} rows · {len(raw_df.columns)} columns detected</div>
          </div>
        </div>""", unsafe_allow_html=True)
        with st.expander("👁 Preview raw data (first 5 rows)"):
            st.dataframe(raw_df.head(), use_container_width=True)
        st.markdown('<hr class="hdiv">', unsafe_allow_html=True)
        st.markdown("""
        <div class="card card-accent">
          <div class="step-num">Step 02</div>
          <div class="card-title">Run Cleaning Process</div>
          <div class="card-desc">Applies all rules, removes invalid rows, builds pivot & 3 charts.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("🧹  Clean Tracker", use_container_width=True, type="primary"):
            with st.spinner("Cleaning data · building pivot · generating charts…"):
                cleaned_df, dropped = clean_tracker(
                    raw_df.copy(), valid_names=valid_names,
                    valid_report_names=valid_reports, valid_regions=valid_regions,
                    valid_statuses=valid_status, threshold=sensitivity)
                pivot_df    = build_pivot(cleaned_df)
                excel_bytes = to_excel_bytes(cleaned_df, pivot_df, tracker_tab_name)
            warn_html = (f'<div class="warn-note">🗑️ {dropped} row(s) removed — '
                         f'empty/unrecognised values in key columns.</div>' if dropped else "")
            st.markdown(f"""
            <div class="success-card">
              ✅ Done — {len(cleaned_df):,} rows cleaned · pivot + 3 charts ready
              {warn_html}
            </div>""", unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                with st.expander("👁 Preview cleaned data (first 5 rows)"):
                    st.dataframe(cleaned_df.head(), use_container_width=True)
            with col2:
                with st.expander("📊 Preview pivot table"):
                    st.dataframe(pivot_df, use_container_width=True, hide_index=True)
            st.markdown('<hr class="hdiv">', unsafe_allow_html=True)
            st.markdown("""
            <div class="card card-accent">
              <div class="step-num">Step 03</div>
              <div class="card-title">Download Cleaned Tracker</div>
              <div class="card-desc">Three sheets: <b>Cleaned Data</b> · <b>Pivot</b> · <b>Charts</b>.</div>
            </div>""", unsafe_allow_html=True)
            st.download_button(
                label=f"⬇️   Download  {output_filename}", data=excel_bytes,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
    except Exception as e:
        st.error(f"❌ Could not process file: {e}")
        st.info("Make sure you upload a valid Excel file (.xlsx or .xls).")
else:
    st.markdown('<div style="text-align:center;padding:52px 0;color:#1E2D48;font-size:0.88rem;">↑ Upload a file above to begin</div>', unsafe_allow_html=True)

st.markdown('<div style="margin-top:56px;text-align:center;font-size:0.68rem;color:#1A2438;font-family:\'Space Mono\',monospace;letter-spacing:1px;">QAR OPS INTERNAL · DAILY TRACKER AUTOMATION</div>', unsafe_allow_html=True)